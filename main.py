from fastapi import FastAPI, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from typing import Optional
import pytesseract
import cv2
import numpy as np
import re
import json
import uuid
import io
from PIL import Image
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


app = FastAPI(title="NotaLens OCR API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def preprocess_image(img_array):
    gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    adaptive = cv2.adaptiveThreshold(
        denoised, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 11, 2
    )
    kernel = np.array([[-1,-1,-1],[-1,9,-1],[-1,-1,-1]])
    sharpened = cv2.filter2D(adaptive, -1, kernel)
    return sharpened

def clean_text(text):
    cleaned = text.strip()
    cleaned = re.sub(r'^[^a-zA-Z0-9]+|[^a-zA-Z0-9]+$', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = re.sub(r'(\d+)\s+(\d{3})', r'\1,\2', cleaned)
    cleaned = re.sub(r'(\d+)\.(\d{3})', r'\1,\2', cleaned)
    return cleaned.strip()

def is_valid_price(text):
    cleaned = re.sub(r'[\s.,]', '', text)
    return bool(re.match(r'^\d{3,8}$', cleaned))

def extract_number_v2(text):
    cleaned = re.sub(r'[^\d]', '', text)
    return int(cleaned) if cleaned and len(cleaned) >= 3 else None

def parse_receipt_v3(ocr_results):
    all_text = [
        clean_text(text)
        for (_, text, conf) in ocr_results
        if conf > 0.3 and clean_text(text)
    ]

    parsed = {
        "raw_text": all_text,
        "nama_toko": None,
        "tanggal": None,
        "total": None,
        "items": [],
        "confidence_score": 0
    }

    for line in all_text:
        if len(line.strip()) > 3 and not is_valid_price(line):
            parsed["nama_toko"] = line.strip().replace("?", "")
            break

    # Tanggal
    bulan_map = {
        'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'mei':5,
        'jun':6,'jul':7,'aug':8,'agu':8,'sep':9,
        'oct':10,'okt':10,'nov':11,'dec':12,'des':12
    }
    date_patterns = [
        r'\d{2}[-/]\d{2}[-/]\d{4}',
        r'\d{4}[-/]\d{2}[-/]\d{2}',
        r'\d{2}[-/]\d{2}[-/]\d{2}',
    ]
    found_date = False
    for line in all_text:
        if found_date: break
        for pattern in date_patterns:
            match = re.search(pattern, line)
            if match:
                parsed["tanggal"] = match.group()
                found_date = True
                break
        match = re.search(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{2,4})', line)
        if match and not found_date:
            day, mon, year = match.groups()
            mon_num = bulan_map.get(mon.lower())
            if mon_num:
                year = year if len(year) == 4 else "20" + year
                parsed["tanggal"] = f"{day.zfill(2)}/{str(mon_num).zfill(2)}/{year}"
                found_date = True

    total_keywords = ['total', 'jumlah', 'grand total', 'bayar', 'tagihan']
    kandidat_total = []
    for i, line in enumerate(all_text):
        if any(kw in line.lower() for kw in total_keywords):
            for j in [i, i-1, i+1]:
                if 0 <= j < len(all_text):
                    num = extract_number_v2(all_text[j])
                    if num and num > 1000:
                        kandidat_total.append(num)
    if kandidat_total:
        parsed["total"] = max(kandidat_total)

    skip_keywords = [
        'total', 'subtotal', 'sub total', 'payment', 'bayar',
        'debit', 'tunai', 'cash', 'thank', 'please', 'closed',
        'check', 'gratis', 'struk', 'jumlah', 'kembalian',
        'change', 'ppn', 'tax', 'disc', 'diskon', 'kembali',
        'terimakasih', 'kritik', 'saran', 'qty', 'link'
    ]
    noise_patterns = [
        r'^Rp\s*[\d,]+$', r'^\d+\s*ml\s*x', r'^lusin\s*x',
        r'^\d+\s*x\s*[\d,]+$', r'^x\s*[\d,]+$', r'^x$',
        r'^Rp$', r'^\d+$',
    ]

    def is_noise(text):
        for pattern in noise_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return True
        return False

    def is_item_name(text):
        if len(text) < 3: return False
        if is_noise(text): return False
        if any(kw in text.lower() for kw in skip_keywords): return False
        if is_valid_price(text): return False
        return True

    numbered_pattern = r'^\d+\s*[.)\-]\s*(.+)$'
    i = 0
    while i < len(all_text):
        line = all_text[i].strip()
        numbered_match = re.match(numbered_pattern, line)
        if numbered_match:
            nama_item = numbered_match.group(1).strip()
            for j in range(i+1, min(i+4, len(all_text))):
                next_line = all_text[j].strip()
                price_match = re.search(r'Rp\s*([\d,]+)|^([\d,]+)$', next_line)
                if price_match:
                    price_str = price_match.group(1) or price_match.group(2)
                    price = extract_number_v2(price_str)
                    if price and 500 < price < (parsed["total"] or 9999999):
                        parsed["items"].append({"nama": nama_item, "harga": price})
                        break
        i += 1

    if not parsed["items"]:
        price_pattern = r'^\d{1,3}[.,]\d{3}$'
        for i, line in enumerate(all_text):
            line = line.strip()
            if not is_item_name(line): continue
            if i + 1 < len(all_text):
                next_line = all_text[i + 1].strip()
                if re.match(price_pattern, next_line):
                    price = extract_number_v2(next_line)
                    if price and 500 < price < (parsed["total"] or 9999999):
                        parsed["items"].append({"nama": line, "harga": price})

    fields_found = sum([
        parsed["nama_toko"] is not None,
        parsed["tanggal"] is not None,
        parsed["total"] is not None,
        len(parsed["items"]) > 0
    ])
    parsed["confidence_score"] = f"{fields_found}/4 fields terdeteksi"
    return parsed

def ocr_pipeline(img_array):
    """OCR pipeline pakai Tesseract — ringan & cepat"""

    # Preprocessing
    processed = preprocess_image(img_array)

    # Convert ke PIL Image
    pil_img = Image.fromarray(processed)

    # Jalankan Tesseract
    # lang='ind+eng' = Bahasa Indonesia + English
    raw_text = pytesseract.image_to_string(
        pil_img,
        lang='ind+eng',
        config='--psm 6'  # PSM 6 = assume single block of text
    )

    # Convert ke format yang sama seperti EasyOCR
    lines = raw_text.split('\n')
    ocr_results = [
        (None, line.strip(), 0.99)
        for line in lines
        if line.strip()
    ]

    return parse_receipt_v3(ocr_results)

def export_to_excel(ocr_jobs, event_id=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pengeluaran"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C0392B")
    center = Alignment(horizontal="center")
    ws.merge_cells("A1:G1")
    ws["A1"] = "LAPORAN PENGELUARAN NOTALENS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.merge_cells("A2:G2")
    filter_label = f"Event: {event_id}" if event_id else "Semua Event"
    ws["A2"] = f"{filter_label} | Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = center
    headers = ["No", "Nama Toko", "Tanggal", "Event ID", "Kategori", "Items", "Total (Rp)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    jobs = list(ocr_jobs.values())
    if event_id:
        jobs = [j for j in jobs if j.get("event_id") == event_id]
    total_keseluruhan = 0
    for idx, job in enumerate(jobs, 1):
        ocr = job.get("ocr_result", {})
        items = ocr.get("items", [])
        items_str = ", ".join([f"{i['nama']} (Rp{i['harga']:,})" for i in items])
        total = ocr.get("total") or 0
        total_keseluruhan += total
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=ocr.get("nama_toko", "-"))
        ws.cell(row=row, column=3, value=ocr.get("tanggal", "-"))
        ws.cell(row=row, column=4, value=job.get("event_id", "-"))
        ws.cell(row=row, column=5, value=job.get("kategori", "-"))
        ws.cell(row=row, column=6, value=items_str)
        ws.cell(row=row, column=7, value=total)
        if idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FADBD8")
    total_row = len(jobs) + 5
    ws.cell(row=total_row, column=6, value="TOTAL KESELURUHAN").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_keseluruhan).font = Font(bold=True)
    column_widths = [5, 20, 15, 15, 15, 50, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.get("/")
def root():
    return {"service": "NotaLens OCR API", "status": "running", "version": "1.0.0"}

@app.post("/ocr")
async def process_receipt(
    file: UploadFile = File(...),
    event_id: Optional[str] = Form(None),
    workspace_id: Optional[str] = Form(None),
    kategori: Optional[str] = Form("organisasi"),
    uploaded_by: Optional[str] = Form(None)
):
    try:
        job_id = str(uuid.uuid4())
        timestamp = datetime.now().isoformat()
        contents = await file.read()
        img = Image.open(io.BytesIO(contents)).convert("RGB")
        img_array = np.array(img)
        ocr_data = ocr_pipeline(img_array)
        result = {
            "job_id": job_id,
            "timestamp": timestamp,
            "filename": file.filename,
            "event_id": event_id,
            "workspace_id": workspace_id,
            "kategori": kategori,
            "uploaded_by": uploaded_by,
            "ocr_result": ocr_data,
            "status": "success"
        }
        ocr_jobs[job_id] = result
        return {"success": True, "job_id": job_id, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/ocr/status/{job_id}")
def get_ocr_status(job_id: str):
    if job_id in ocr_jobs:
        return {"success": True, "job_id": job_id, "data": ocr_jobs[job_id]}
    return {"success": False, "error": f"Job ID '{job_id}' tidak ditemukan"}

@app.get("/ocr/history")
def get_history(event_id: Optional[str] = None):
    if event_id:
        filtered = {k: v for k, v in ocr_jobs.items() if v.get("event_id") == event_id}
        return {"success": True, "count": len(filtered), "data": filtered}
    return {"success": True, "count": len(ocr_jobs), "data": ocr_jobs}

@app.get("/ocr/summary")
def get_summary(event_id: Optional[str] = None):
    jobs = list(ocr_jobs.values())
    if event_id:
        jobs = [j for j in jobs if j.get("event_id") == event_id]
    total_pengeluaran = sum(j.get("ocr_result", {}).get("total") or 0 for j in jobs)
    per_toko = {}
    for j in jobs:
        toko = j.get("ocr_result", {}).get("nama_toko", "Unknown")
        total = j.get("ocr_result", {}).get("total") or 0
        per_toko[toko] = per_toko.get(toko, 0) + total
    per_kategori = {}
    for j in jobs:
        kat = j.get("kategori", "unknown")
        total = j.get("ocr_result", {}).get("total") or 0
        per_kategori[kat] = per_kategori.get(kat, 0) + total
    return {
        "success": True,
        "event_id": event_id,
        "summary": {
            "total_scan": len(jobs),
            "total_pengeluaran": total_pengeluaran,
            "per_toko": per_toko,
            "per_kategori": per_kategori
        }
    }

@app.get("/ocr/export")
def export_excel(event_id: Optional[str] = None):
    try:
        excel_file = export_to_excel(ocr_jobs, event_id)
        filename = f"laporan_{event_id or 'semua'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return {"success": False, "error": str(e)}